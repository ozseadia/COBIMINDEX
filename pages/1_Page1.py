# -*- coding: utf-8 -*-
"""
Created on Sat Aug 12 12:37:32 2023

@author: OzSea
"""
import pandas as pd
import os
import time
from datetime import timedelta ,datetime 
import numpy as np
import streamlit as st
import openpyxl
import DeshbordAPP as De
from PIL import Image

global path_svg
#path_svg=(r'G:\Oz\fiveer\Dani_Velinchick\KrohnApp\python_codes\pages\SVG\OpenScreenLogo.svg')
st.set_page_config(page_title="Patient details settings", page_icon=None, layout="wide", initial_sidebar_state="auto", menu_items=None)


#rootPath=r'G:\Oz\fiveer\Dani_Velinchick\KrohnApp\python_codes'
dirname = os.path.dirname(__file__)
filename=os.path.join(dirname,'../Data/acount and passwords.xlsx')
Table_acounts=pd.read_excel(filename)

#SetUp 
Cor_options=['...','Noa Shultz','Hila Askayo','Ganit Goren']
S_options=['...',
           'Naama Peled-Ironi',
           'Milca Hanukoglu',
           'Helen Israel',
           'Adi Vilensky',
           'Noa Schultz',
           'Oneg Kabizon',
           'Shachar Michael',
           'Zohar Peled-Zinger',
           'Ganit Goren'
           ]


def excelupdate(PatientNumber,SW,Cordinator,ind,T1,T2,T3,T4,T5):
    #read the existing sheets so that openpyxl won't create a new one later
    wb = openpyxl.load_workbook(filename)
    ws=wb['Sheet1']
    if PatientNumber:
        ws.cell(row=ind, column=4).value=PatientNumber
    if not (SW=='...'):    
        ws.cell(row=ind, column=5).value=SW
    if not (Cordinator=='...'):
        ws.cell(row=ind, column=6).value=Cordinator
    ws.cell(row=ind, column=7).value=T1.strftime('%Y-%m-%d')
    ws.cell(row=ind, column=8).value=T2.strftime('%Y-%m-%d')
    ws.cell(row=ind, column=9).value=T3.strftime('%Y-%m-%d')
    ws.cell(row=ind, column=10).value=T4.strftime('%Y-%m-%d')
    ws.cell(row=ind, column=11).value=T5.strftime('%Y-%m-%d')
    wb.save(filename)
    # writer = pd.ExcelWriter(filename, engine='openpyxl') 
    # writer.book = book
    # writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    
    # #update without overwrites
    # # update=pd.DataFrame({'A':[3,4],
    # #                      'B':[4,5]}, index=(pd.date_range('2004-04-30', 
    # #                                                      periods=2,
    # #                                                      freq='M').strftime('%Y-%m-%d')))
    # update=pd.DataFrame({'Social worker':list(Table_acounts.iloc[ind]['Social worker']),
    #                      'Coordinator':list(Table_acounts.iloc[ind]['Coordinator'])})
    # update.to_excel(writer, "Sheet1", startrow=ind, startcol=4)
    
    # writer.save()

def excelupdate1(Status,ind):
    if Status:
        wb = openpyxl.load_workbook(filename)
        ws=wb['Sheet1']
        ws.cell(row=ind, column=12).value=Status
        wb.save(filename)
        



def Refrash(ind,PP):
    Table_acounts=pd.read_excel(filename)
    PP.dataframe(Table_acounts.iloc[ind,[0,2,3,4,5,6,7,8,9,10,11]])
    

De.render_svg(De.read_svg())
st.title('Patient details settings')
options=list(Table_acounts.acount)
options.insert(0,'.....')
NAME=st.sidebar.selectbox('Select Patient', options)
if not(NAME=='.....'):
    ind=Table_acounts.index[Table_acounts['acount']==int(NAME)]
    PP=st.dataframe(Table_acounts.iloc[ind,[0,2,3,4,5,6,7,8,9,10,11]])
    col1, col2 ,col3 = st.columns(3)
    with col1:
        Cordinator=st.selectbox('Assigned Coordinator',Cor_options)
        SW=st.selectbox('Assigned Therapist',S_options)
        PatienNamber=st.text_input('Patient ID')
    if (Table_acounts.iloc[ind,[6,7,8,9,10]]).isnull().values.any():
        with col2:
        #today = datetime.datetime.now().date()
            T1=st.date_input('T1 start date',datetime.now().date())
            T2=st.date_input('T2 start date',datetime.now().date()+timedelta(days=90))
            T3=st.date_input('T3 start date',datetime.now().date()+timedelta(days=90*2))
        with col3:    
            T4=st.date_input('T4 start date',datetime.now().date()+timedelta(days=90*3))
            T5=st.date_input('T5 start date',datetime.now().date()+timedelta(days=90*4)) 
    else:
        with col2:            
            T1=st.date_input('T1 start date',datetime.strptime(Table_acounts['T1'][ind].iloc[0],"%Y-%m-%d").date())
            T2=st.date_input('T2 start date',datetime.strptime(Table_acounts['T2'][ind].iloc[0],"%Y-%m-%d").date())
            T3=st.date_input('T3 start date',datetime.strptime(Table_acounts['T3'][ind].iloc[0],"%Y-%m-%d").date())
        with col3:
            T4=st.date_input('T4 start date',datetime.strptime(Table_acounts['T4'][ind].iloc[0],"%Y-%m-%d").date())
            T5=st.date_input('T5 start date',datetime.strptime(Table_acounts['T5'][ind].iloc[0],"%Y-%m-%d").date())
            
    if st.button('Update'):
        excelupdate(PatienNamber,SW,Cordinator,int(ind[0])+2,T1,T2,T3,T4,T5)
        time.sleep(0.5)
        Refrash(ind,PP)
        
    col4, col5 ,col6 = st.columns(3)
    with col4:    
        Status=st.selectbox('Patient Status',['Working_ok','Need_attention','blocked'])
        if st.button('Submit status'):
            excelupdate1(Status,int(ind[0]+2))
            Refrash(ind,PP)
            
    with open(filename, 'rb') as my_file:
        st.download_button(label = 'Download', data = my_file, file_name = filename, mime = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')   
        
