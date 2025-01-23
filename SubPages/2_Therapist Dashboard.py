# -*- coding: utf-8 -*-
"""
Created on Sun Aug 13 08:31:49 2023

@author: OzSea
"""

import os
import time
from datetime import timedelta ,datetime 
import openpyxl
import streamlit as st
import DataBase as DB
import pandas as pd
import numpy as np
import altair as alt
import plotly.express as px 
import plotly.graph_objects as go
from Functions import render_svg , read_svg,load_data , Table1, PlotyCandlestick #as De

#st.set_page_config(page_title="Patients Dashboard", page_icon=None, layout="wide", initial_sidebar_state="auto", menu_items=None)
if 'page_config' not in st.session_state:
    st.session_state['page_config'] = st.set_page_config(page_title="Patient details settings", page_icon=None, layout="wide", initial_sidebar_state="auto", menu_items=None)


if 'LogIn_Aprove' not in st.session_state:
    st.session_state['LogIn_Aprove'] = []
if 'Patients' not in st.session_state:
    st.session_state['Patients'] = 1

if 'loaddata' not in st.session_state:    #st.session_state.loaded_data = load_data()
    T,V,date,userid,ActiveUsers_id=load_data(temp=1)
    st.session_state['loaddata']=1
    st.session_state['T']=T
    st.session_state['V']=V
    st.session_state['date']=date
    st.session_state['userid']=userid
    st.session_state['ActiveUsers_id']=ActiveUsers_id

T=st.session_state['T']
V=st.session_state['V']
date=st.session_state['date']
userid=st.session_state['userid']
ActiveUsers_id=st.session_state['ActiveUsers_id'] 
   
#st.set_page_config(page_title="Patient details settings", page_icon=None, layout="wide", initial_sidebar_state="auto", menu_items=None)
#st.set_page_config(layout="wide") 
#path_svg=(r'G:\Oz\fiveer\Dani_Velinchick\KrohnApp\python_codes\pages\SVG\OpenScreenLogo.svg')
render_svg(read_svg())
S_options=['...',
           'Naama Peled-Ironi',
           'Milca Hanukoglu',
           'Helen Israel',
           'Adi Vilensky',
           'Noa Schultz',
           'Oneg Kabizon',
           'Shachar Michael',
           'Zohar Peled-Zinger',
           'Noa Shultz','Hila Askayo','Ganit Goren','General']

@st.cache_data()
def EnableAll():
    st.session_state['Patients']=1

SW=st.sidebar.selectbox('Assigned Coordinator/Therapist',S_options,on_change=EnableAll())
#SW=st.sidebar.selectbox('Assigned Coordinator/Therapist',[str(st.session_state['LogIn_Aprove'][0][0])],on_change=EnableAll())
#SW=st.sidebar.selectbox('Assigned Coordinator/Therapist',S_options)
#time.sleep(1)
#EnableAll(SW)
# rootPath=r'G:\Oz\fiveer\Dani_Velinchick\KrohnApp\python_codes'
# filename=os.path.join(rootPath,'Data/acount and passwords.xlsx')
dirname = os.path.dirname(__file__)
filename=os.path.join(dirname,'../Data/acount and passwords.xlsx')
Table_acounts=pd.read_excel(filename)


def highlight_max(x, color):
    return np.where(x <=-10, "background-color: {color}", None)

def highlight_cols(s):
    color = 'red' if s < 0 else ''
    return 'background-color: % s' % color

def highlight_cols1(s):
    color = 'red' if s > 2 else ''
    return 'background-color: % s' % color
def apply_formatting(col):
    if col.name == 'SUDS' or col.name == 'Pain' or col.name == 'Fatigue':
        a=[]
        for c in col.values:
            temp=True
            for i in range(len(c)):
                temp=temp*(c[i][0]-c[i][1] < 0) 
            a.append(temp)
        #return ['background-color :red' if v  else 'color: blue'    for sub in a for v in sub]
        return ['background-color :red' if v  else 'background-color: green'    for v in a]
    elif (col.name == 'Well-being') :
        a=[]
        for c in col.values:
            temp=True
            for i in range(len(c)):
                temp=temp*(c[i][0]-c[i][1] > 0) 
            a.append(temp)
        #return ['background-color :red' if v  else 'color: blue'    for sub in a for v in sub]
        return ['background-color :red' if v  else 'background-color: green'    for v in a]
        
        
    else:    
        return ['color: black ' for c in col.values]
def Chart_data(Table2):
    Temp=pd.DataFrame(columns=['Value'])
    Temp['Value']=pd.concat([Table2['sud power'],Table2['vas power'],Table2['fat power'],Table2['well power']])
    df1=pd.DataFrame(columns=['Ind','Date'])
    df1['Ind']=pd.DataFrame(['sud']*len(Table2)+['vas']*len(Table2)+['fat']*len(Table2)+['well']*len(Table2),
                     index=Temp['Value'].index)
    df1['Date']=list(Temp.index)
    C_data=pd.concat([Temp['Value'],df1['Ind'],df1['Date']],axis="columns")
    C=alt.Chart(C_data).mark_bar().encode(
    x='Ind:O',
    y='Value:Q',
    color='Ind:N',
    column='Date:N')
    return(C)

def Ploty(Table2,Ind):
    plot=px.bar(x=Table2.index,y=Table2[Ind])
    #plot['layout']['xaxis']['autorange'] = "reversed"
    return plot

def PlotyMulty(Table2,Ind):
    plot=px.line(x=Table2.index,y=Table2[Ind])
    return(plot)
    pass

# def PlotyCandlestick(Table2,Ind):
#     #plot=px.bar(x=Table2.index,y=Table2[Ind])
#     Topen=Ind+"1"
#     Tclose=Ind+"2"
    
#     if Ind=='well':
#         plot = go.Figure(data=[go.Candlestick(x=Table2.index,
#                     open=Table2[Topen],
#                     high=Table2[[Topen,Tclose]].max(axis=1),
#                     low=Table2[[Topen,Tclose]].min(axis=1),
#                     close=Table2[Tclose],
#                     increasing_line_color= 'green', decreasing_line_color= 'red')])
#     else:
#         plot = go.Figure(data=[go.Candlestick(x=Table2.index,
#                     open=Table2[Topen],
#                     high=Table2[[Topen,Tclose]].max(axis=1),
#                     low=Table2[[Topen,Tclose]].min(axis=1),
#                     close=Table2[Tclose],
#                     increasing_line_color= 'red', decreasing_line_color= 'green')])
        
#     plot.update_layout(xaxis_rangeslider_visible=False)
#     #plot['layout']['xaxis']['autorange'] = "reversed"
#     return plot
    
def ComplinesTable(TABLE):
    ind=list(range(0,14))
    return(TABLE.iloc[ind])

def IndexTable(TABLE):
    ind=list(range(14,len(TABLE)))
    return(TABLE.iloc[ind])
def Extract_Patien_List():
    a=Table_acounts.query('Therapist==@SW or Coordinator==@SW')
    return a['acount']
def Extract_Patien_ListAppId():
    a=Table_acounts.query('Therapist==@SW or Coordinator==@SW')
    return a
def ReplaceKeys(T,PatientList):
    for i in PatientList.index:
        T.columns = T.columns.str.replace(str(PatientList['acount'][i]), str(PatientList['Patient_ID'][i]))
    return(T)
def ConvertPatienID2Acount(PatienID,PatientList):
    return(PatientList['acount'][PatientList['Patient_ID']==PatientID].iloc[0])

def submit_text(message,ind):
    if not(message==""):
        wb = openpyxl.load_workbook(filename)
        ws=wb['Sheet1']
        #Text= str(ws.cell(row=ind, column=13).value)+"|"+"\n"
        Text=str(ws.cell(row=ind, column=13).value)
        if not(str(ws.cell(row=ind, column=13).value)=='None'):
            ws.cell(row=ind, column=13).value =str(ws.cell(row=ind, column=13).value)+"\n"+str(message)+str(datetime.now())[:-10]+"|"
        else:
            ws.cell(row=ind, column=13).value =str(message)+str(datetime.now())[:-10]+"|"
        wb.save(filename)
        
def Refrash(ind,PP):
    Table_acounts=pd.read_excel(filename)
    PP.dataframe(Table_acounts.iloc[ind,[12]])        

@st.cache_data(ttl=3600)
def load_data(temp):
    
    T,V,date,userid,ActiveUsers_id=DB.start()
    return (T,V,date,userid,ActiveUsers_id)    

# @st.cache_data(ttl=3600)
# def Table1(V,date,ActiveUsers_id,TypeSession='Morning',):
#     TABLE=DB.Table1(V,date,ActiveUsers_id,'Morning','.....')
#     return (TABLE)
   
st.title('Dashboard')
#T,V,date,userid,ActiveUsers_id=DB.start()

#T,V,date,userid,ActiveUsers_id=De.load_data(temp=1)

if not(SW=='...'):
    PatientList=Extract_Patien_ListAppId()
    #PatientList=Extract_Patien_List()
    ids_List=[]
    for temp in PatientList['acount']:
        ids_List.append(DB.Convert_acount2id(V,temp))
    #TABLE=DB.Table1(V,date,ids_List,'Morning','.....')
    TABLE=Table1(V,date,ids_List,'Morning')
    
#TABLE.style.apply(highlight_max, color='red')
    placeholder1 = st.empty()
    if st.session_state['Patients']>=1:
        with placeholder1.container():
            st.subheader(" :woman-running: :runner: Patients' Adherence Table")
            Tc=ComplinesTable(TABLE)
            Tc=ReplaceKeys(Tc,PatientList)
            
            #Tc = Tc.append([PatientList['Patient_Status']], ignore_index = False)
            #print(type(PatientList['Patient_Status']))
            #new_record = pd.DataFrame([PatientList['Patient_Status']], columns=Tc.columns)
            #Tc = pd.concat([new_record, Tc], ignore_index=False)
            #Tc=pd.concat(pd.DataFrame(PatientList['Patient_Status']),Tc,ignore_index = False)
            if not Tc.empty:
                Tc.loc[-1]=list(PatientList['Patient_Status'])    
                Tc.rename({Tc.index[-1]:'Patient_Status' }, inplace=True)
                Tc = Tc.reindex(np.roll(Tc.index, shift=1))
            st.dataframe(Tc.style.applymap(highlight_cols1,
                                            subset = pd.IndexSlice[['Lag days in current Level','Total Lag days'],:]).format(precision=0))
            st.subheader("Patients' Index table")
            col1,col2 = st.columns([1,1])
            with col1:
                TypeSession=st.selectbox('Select Morning or Evenining :sun_with_face:/:first_quarter_moon_with_face:',['Morning','Evening'])
            #TABLE=DB.Table1(V,date,ids_List,TypeSession,'.....')
            TABLE=Table1(V,date,ids_List,TypeSession)
            Ti=IndexTable(TABLE)
            Ti=ReplaceKeys(Ti,PatientList)
            st.dataframe(Ti.style.applymap(highlight_cols,
                                            ).format(precision=0))
    

    options=list(PatientList['Patient_ID'])
    options.insert(0,'.....')
    PatientID=st.sidebar.selectbox('Select Patient', options)
    if not (PatientID==options[0]):
        st.session_state['Patients']=0
        NAME=str(ConvertPatienID2Acount(PatientID,PatientList))
        ind=Table_acounts.index[Table_acounts['acount']==int(NAME)]
        placeholder1.empty()
        time.sleep(0.01)
        with placeholder1.container():
            
            Table3,Level,GroupType=DB.Patient_Records(T,V,NAME)
            st.title(':chart_with_downwards_trend: :chart_with_upwards_trend: Patient '+ PatientID + ' Dashboard')
            st.subheader('Level :'+str(int(Level)))
            st.subheader('Group Type: '+ GroupType)
            st.dataframe(Table3.iloc[::-1].style.apply(apply_formatting))
            
            #Message=st.text_area("insert free text",value="",key='text2')
            with st.form("insert free text",clear_on_submit=True):
                PP=st.dataframe((Table_acounts.iloc[ind,[12]]),use_container_width=True,hide_index=True)
                #Message = st.text_input("insert free text:")
                Message=st.text_area("insert free text",value="",key='text2')
                submitted = st.form_submit_button("Submit")
                if submitted:
                    submit_text(Message,int(ind[0]+2))
                    Refrash(ind,PP)
                    #Message=[]
            # if st.button('Submit text'):
            #     submit_text(Message,int(ind[0]+2))
            #     Refrash(ind,PP)
            st.title(':clipboard: Patient '+PatientID +' Exercised techniques')
            weeks=DB.weeks_from_Start(V,NAME,date)
            #st.subheader(str(int(Tc[PatientID].iloc[9]))+' Weeks from start')
            st.subheader(str(weeks)+' Weeks from start')
            
            Table4=DB.technics(V,NAME,date)
            #df1=Table3.iloc[:, 2:4]
            st.dataframe(Table4.iloc[:,0:5].set_index('technic number').style.format(precision=0))    
            #col1,col2= st.columns([1,3])
            #with col1:
            #    TypeSession=st.selectbox('Please Select Morning or Evenining :sun_with_face:/:first_quarter_moon_with_face:',['Morning','Evening'])
            #    time.sleep(1)
            TypeSession='Morning'    
            Table02=DB.userData(V,date,NAME,TypeSession)
            Table2=Table02[0:-2]
            #Table3=DB.Patient_Records(T,V,NAME)
            #C=Chart_data(Table2)
            
            st.subheader(':chart_with_downwards_trend: :chart_with_upwards_trend: Patient '+ PatientID + ' ' + TypeSession+ ' indexes results')
            #st.dataframe(Table2.style.applymap(highlight_cols).format(precision=0))
            #st.dataframe(Table3.style.apply(apply_formatting))
            #st.altair_chart(C)
            col4,col5,col6 = st.columns(3)
            height=300
            with col4:
                try:
                    #col4.empty()
                    #st.subheader('SUD')
                    plot1=PlotyCandlestick(Table2,'sud')
                    plot1.update_layout(title_text='SUD',title_x=0.4,height=height)
                #plot=De.PlotyMulty(Table2)
                    st.plotly_chart(plot1,use_container_width=True)
                #st.bar_chart(Table2['sud power'])
                except:
                    pass
            with col5:
                try:
                    #col5.empty()
                    #st.subheader('VAS')
                    plot2=PlotyCandlestick(Table2,'vas')
                    plot2.update_layout(title_text='VAS',title_x=0.4,height=height)
                    st.plotly_chart(plot2,use_container_width=True)
                #st.bar_chart(Table2['vas power'])
                except:
                    pass
            with col6:
                try:
                    #col6.empty()
                    #st.subheader('Fatigue')
                    plot3=PlotyCandlestick(Table2,'fat')
                    plot3.update_layout(title_text='Fatigue',title_x=0.4,height=height)
                    st.plotly_chart(plot3,use_container_width=True)
                    #st.bar_chart(Table2['fat power'])
                except:
                    pass
                
            TypeSession='Evening'    
            Table02=DB.userData(V,date,NAME,TypeSession)
            Table2=Table02[0:-2]
            #Table3=DB.Patient_Records(T,V,NAME)
            #C=Chart_data(Table2)
            
            st.subheader(':chart_with_downwards_trend: :chart_with_upwards_trend: Patient '+ PatientID + ' ' + TypeSession+ ' indexes results')
            #st.dataframe(Table2.style.applymap(highlight_cols).format(precision=0))
            #st.dataframe(Table3.style.apply(apply_formatting))
            #st.altair_chart(C)
            col4,col5,col6 = st.columns(3)
            with col4:
                try:
                    #col4.st.empty()
                    #st.subheader('SUD')
                    plot1=PlotyCandlestick(Table2,'sud')
                    plot1.update_layout(title_text='SUD',title_x=0.4,height=height)
                #plot=De.PlotyMulty(Table2)
                    st.plotly_chart(plot1,use_container_width=True)
                #st.bar_chart(Table2['sud power'])
                except:
                    pass
            with col5:
                try:
                    #col5.empty()
                    #st.subheader('VAS')
                    plot2=PlotyCandlestick(Table2,'vas')
                    plot2.update_layout(title_text='VAS',title_x=0.4,height=height)
                    st.plotly_chart(plot2,use_container_width=True)
                #st.bar_chart(Table2['vas power'])
                except:
                    pass
            with col6:
                try:
                    #col6.empty()
                    #st.subheader('Fatigue')
                    plot3=PlotyCandlestick(Table2,'fat')
                    plot3.update_layout(title_text='Fatigue',title_x=0.4,height=height)
                    st.plotly_chart(plot3,use_container_width=True)
                    #st.bar_chart(Table2['fat power'])
                except:
                    pass
            #st.subheader('Wellbeing')
            #plot4=De.PlotyCandlestick(Table2,'well')
            #plot4.update_layout(title_text='Wellbeing',title_x=0.5)
            #st.plotly_chart(plot4)
            #st.bar_chart(Table2['well power'])
    
            #st.bar_chart(pd.DataFrame(chart_data))
            #userID=str(V['App_user'].id[V['App_user'].username==int(NAME)].values)[1:-1]
            
            #st.text(userID)
    else :
        st.session_state['Patients']+=1
        time.sleep(1)
        if st.session_state['Patients']==1 :
            #st.experimental_rerun()
            st.rerun()
else:
    st.subheader('Please choose Therapist / Assigned Coordinator’ s name')
            